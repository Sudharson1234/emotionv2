"""
Report Export Module
Handles exporting chat data and emotion analytics to Excel
"""

import logging
import os
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlsxwriter

logger = logging.getLogger(__name__)

# Color scheme for emotions
EMOTION_COLORS = {
    'joy': 'FFD700',      # Gold
    'sadness': '4169E1',  # Royal Blue
    'anger': 'FF4500',    # Orange Red
    'fear': '8B008B',     # Dark Magenta
    'disgust': '228B22',  # Forest Green
    'surprise': 'FF1493', # Deep Pink
    'neutral': 'A9A9A9',  # Dark Gray
}

def export_chat_to_excel(chats, filename=None, domain_name='EmotiChat'):
    """
    Export chat history to Excel file with emotion analysis.
    
    Args:
        chats (list): List of chat messages with emotion data
        filename (str): Output filename (optional)
        domain_name (str): Domain/organization name
        
    Returns:
        BytesIO: Excel file as bytes
    """
    try:
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"EmotiChat_Report_{timestamp}.xlsx"
        
        # Create BytesIO object
        output = BytesIO()
        
        # Create workbook using xlsxwriter
        workbook = xlsxwriter.Workbook(output, {'remove_timezone': True})
        
        # Create worksheets
        worksheet = workbook.add_worksheet('Chat History')
        summary_ws = workbook.add_worksheet('Summary')
        emotion_ws = workbook.add_worksheet('Emotion Analysis')
        
        # Define formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#1a202c',
            'font_color': '#ffffff',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True,
        })
        
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'bg_color': '#00d4ff',
            'font_color': '#1a202c',
            'align': 'center',
            'valign': 'vcenter',
        })
        
        info_format = workbook.add_format({
            'bold': True,
            'font_size': 11,
            'bg_color': '#e8e8e8',
        })
        
        user_msg_format = workbook.add_format({
            'bg_color': '#e3f2fd',
            'border': 1,
            'align': 'left',
            'valign': 'top',
            'text_wrap': True,
        })
        
        ai_msg_format = workbook.add_format({
            'bg_color': '#f3e5f5',
            'border': 1,
            'align': 'left',
            'valign': 'top',
            'text_wrap': True,
        })
        
        emotion_format = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
        })
        
        datetime_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': 'yyyy-mm-dd hh:mm:ss',
        })
        
        # ===== MAIN CHAT HISTORY WORKSHEET =====
        worksheet.set_column('A:A', 12)  # Date/Time
        worksheet.set_column('B:B', 20)  # User
        worksheet.set_column('C:C', 35)  # User Message
        worksheet.set_column('D:D', 35)  # AI Response
        worksheet.set_column('E:E', 15)  # Emotion
        worksheet.set_column('F:F', 12)  # Confidence
        worksheet.set_column('G:G', 20)  # Domain
        
        # Title row
        # xlsxwriter uses merge_range instead of merge_cells
        worksheet.merge_range('A1:G1', 'EmotiChat - Chat History Report', title_format)
        worksheet.set_row(0, 30)
        
        # Info rows
        worksheet.write('A2', f'Report Generated:', info_format)
        worksheet.write('B2', datetime.now().strftime('%Y-%m-%d %I:%M:%S %p'), info_format)
        worksheet.write('A3', f'Domain:', info_format)
        worksheet.write('B3', domain_name, info_format)
        
        # Headers
        headers = ['Date/Time', 'User', 'User Message', 'AI Response', 'Emotion', 'Confidence %', 'Domain']
        for col, header in enumerate(headers, 1):
            worksheet.write(3, col-1, header, header_format)
        worksheet.set_row(3, 25)
        
        # Data rows
        row = 4
        emotion_stats = {}
        
        for chat in chats:
            try:
                # Extract data
                timestamp_str = chat.get('timestamp', '')
                user = chat.get('username', 'Anonymous')
                user_msg = chat.get('user_message', '')
                ai_resp = chat.get('ai_response', '')
                emotion = chat.get('detected_emotion', chat.get('detected_text_emotion', 'neutral'))
                
                # Get confidence
                emotion_score = chat.get('emotion_score', 0)
                confidence = round(emotion_score * 100, 2) if emotion_score else 0
                
                # Write row
                worksheet.write(row, 0, timestamp_str, datetime_format)
                worksheet.write(row, 1, user, user_msg_format)
                worksheet.write(row, 2, str(user_msg)[:500], user_msg_format)
                worksheet.write(row, 3, str(ai_resp)[:500] if ai_resp else '', ai_msg_format)
                worksheet.write(row, 4, emotion, emotion_format)
                worksheet.write(row, 5, confidence, emotion_format)
                worksheet.write(row, 6, domain_name, user_msg_format)
                
                worksheet.set_row(row, 40)
                
                # Track emotion statistics
                emotion_stats[emotion] = emotion_stats.get(emotion, 0) + 1
                
                row += 1
            except Exception as e:
                logger.warning(f"Error writing chat row: {e}")
                continue
        
        # ===== SUMMARY WORKSHEET =====
        summary_ws.set_column('A:A', 20)
        summary_ws.set_column('B:B', 15)
        summary_ws.set_column('C:C', 15)
        
        # merge header for summary sheet
        summary_ws.merge_range('A1:C1', 'EmotiChat - Summary Report', title_format)
        summary_ws.set_row(0, 30)
        
        summary_ws.write('A2', 'Report Metadata', workbook.add_format({'bold': True, 'font_size': 12}))
        summary_ws.write('A3', 'Domain Name:', info_format)
        summary_ws.write('B3', domain_name)
        
        summary_ws.write('A4', 'Report Generated:', info_format)
        summary_ws.write('B4', datetime.now().strftime('%Y-%m-%d %I:%M:%S %p'))
        
        summary_ws.write('A5', 'Total Messages:', info_format)
        summary_ws.write('B5', len(chats))
        
        summary_ws.write('A7', 'Emotion Distribution', workbook.add_format({'bold': True, 'font_size': 12}))
        summary_ws.write('A8', 'Emotion', header_format)
        summary_ws.write('B8', 'Count', header_format)
        summary_ws.write('C8', 'Percentage', header_format)
        
        row = 9
        for emotion, count in sorted(emotion_stats.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / len(chats) * 100) if len(chats) > 0 else 0
            summary_ws.write(row, 0, emotion, emotion_format)
            summary_ws.write(row, 1, count, emotion_format)
            summary_ws.write(row, 2, f'{percentage:.1f}%', emotion_format)
            row += 1
        
        # ===== EMOTION ANALYSIS WORKSHEET =====
        emotion_ws.set_column('A:A', 20)
        emotion_ws.set_column('B:B', 35)
        emotion_ws.set_column('C:C', 12)
        emotion_ws.set_column('D:D', 20)
        
        # merge header for emotion analysis
        emotion_ws.merge_range('A1:D1', 'EmotiChat - Emotion Timeline Analysis', title_format)
        emotion_ws.set_row(0, 30)
        
        emotion_ws.write('A2', 'Date/Time', header_format)
        emotion_ws.write('B2', 'Message Preview', header_format)
        emotion_ws.write('C2', 'Emotion', header_format)
        emotion_ws.write('D2', 'Confidence', header_format)
        emotion_ws.set_row(1, 20)
        
        row = 2
        for chat in chats:
            try:
                timestamp_str = chat.get('timestamp', '')
                user_msg = chat.get('user_message', '')
                emotion = chat.get('detected_emotion', chat.get('detected_text_emotion', 'neutral'))
                emotion_score = chat.get('emotion_score', 0)
                confidence = round(emotion_score * 100, 2) if emotion_score else 0
                
                emotion_ws.write(row, 0, timestamp_str, datetime_format)
                emotion_ws.write(row, 1, str(user_msg)[:100], user_msg_format)
                emotion_ws.write(row, 2, emotion, emotion_format)
                emotion_ws.write(row, 3, f'{confidence}%', emotion_format)
                
                row += 1
            except Exception as e:
                logger.warning(f"Error writing emotion row: {e}")
                continue
        
        # Close workbook
        workbook.close()
        
        # Reset BytesIO position to beginning
        output.seek(0)
        
        logger.info(f"Successfully exported {len(chats)} chats to Excel")
        
        return output
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise


def export_emotion_report(emotion_data, filename=None, domain_name='EmotiChat'):
    """
    Export emotion statistics report to Excel.
    
    Args:
        emotion_data (dict): Emotion statistics data
        filename (str): Output filename (optional)
        domain_name (str): Domain/organization name
        
    Returns:
        BytesIO: Excel file as bytes
    """
    try:
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"EmotiChat_Emotion_Report_{timestamp}.xlsx"
        
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'remove_timezone': True})
        worksheet = workbook.add_worksheet('Emotion Report')
        
        # Define formats
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'bg_color': '#00d4ff',
            'font_color': '#1a202c',
            'align': 'center',
            'valign': 'vcenter',
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#1a202c',
            'font_color': '#ffffff',
            'border': 1,
            'align': 'center',
        })
        
        data_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'num_format': '0.0%',
        })
        
        # Set column widths
        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 15)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 20)
        
        # Title
        worksheet.merge_range('A1:D1', f'EmotiChat - Emotion Analytics Report', title_format)
        worksheet.set_row(0, 30)
        
        # Report info
        worksheet.write('A2', 'Report Generated:', workbook.add_format({'bold': True}))
        worksheet.write('B2', datetime.now().strftime('%Y-%m-%d %I:%M:%S %p'))
        worksheet.write('A3', 'Domain:', workbook.add_format({'bold': True}))
        worksheet.write('B3', domain_name)
        
        # Headers
        worksheet.write('A5', 'Emotion', header_format)
        worksheet.write('B5', 'Count', header_format)
        worksheet.write('C5', 'Percentage', header_format)
        worksheet.write('D5', 'Average Confidence', header_format)
        worksheet.set_row(4, 20)
        
        # Data
        row = 5
        total_count = sum(emotion_data.values()) if emotion_data else 0
        
        for emotion, count in sorted(emotion_data.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total_count * 100) if total_count > 0 else 0
            worksheet.write(row, 0, emotion)
            worksheet.write(row, 1, count)
            worksheet.write(row, 2, percentage / 100, data_format)
            worksheet.write(row, 3, 0.0, data_format)  # Placeholder for average confidence
            row += 1
        
        workbook.close()
        output.seek(0)
        
        logger.info("Successfully exported emotion report to Excel")
        
        return output
        
    except Exception as e:
        logger.error(f"Error exporting emotion report: {e}")
        raise


if __name__ == "__main__":
    # Test data
    test_chats = [
        {
            'timestamp': '2024-02-17 10:30:00',
            'username': 'User1',
            'user_message': 'I am very happy today!',
            'ai_response': 'That is great news!',
            'detected_emotion': 'joy',
            'emotion_score': 0.92,
        },
        {
            'timestamp': '2024-02-17 10:35:00',
            'username': 'User2',
            'user_message': 'I am feeling sad',
            'ai_response': 'I understand your feelings',
            'detected_emotion': 'sadness',
            'emotion_score': 0.87,
        },
    ]
    
    excel_file = export_chat_to_excel(test_chats, domain_name='Test Domain')
    with open('test_report.xlsx', 'wb') as f:
        f.write(excel_file.getvalue())
    
    print("Test report created: test_report.xlsx")
